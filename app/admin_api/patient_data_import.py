from werkzeug.datastructures import FileStorage
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from web_errors import WebError
from dataclasses import dataclass
from typing import Iterable
from clinics.data_access import get_most_common_clinic
from patients.data_access import patient_from_key_data, add_patient
from patients.patient import Patient
from visits.data_access import first_visit_by_patient_and_date, add_visit
from visits.visit import Visit
from events.data_access import clear_all_events, add_event
from events.event import Event
import uuid
from language_strings.language_string import LanguageString
from util import as_string
from datetime import date, timedelta, datetime
import itertools
import json
from config import DEFAULT_PROVIDER_ID_FOR_IMPORT
import pandas as pd
import dateutil

COLUMNS = [
    # 'camp',
    'visit_date',
    'visit_type',
    'doctor',
    'medical_record_num',
    'first_name',
    'surname',
    'date_of_birth',
    'age',
    'gender',
    'origin',
    # 'hometown',
    # 'home_country',
    'municipality',
    'local_id',
    'address'
    'phone',
    'email',
    'attention_datetime',
    'attending_resources',
    'educational_status',
    'religion',
    'marital_status',
    'occupation',
    'mother_name',
    'father_name',
    'delivery_place',
    'delivery_datetime',
    'gestational_age',
    'delivery_care',
    'delivery_via',
    'presentation',
    'birthing_events',

    # # medical hx
    # 'allergies',
    # 'surgery_hx',
    # 'chronic_conditions',
    # 'current_medications',
    # 'vaccinations',
    # # complaint
    # 'complaint',
    # # vitals
    # 'heart_rate',
    # 'blood_pressure',
    # 'sats',
    # 'temp',
    # 'respiratory_rate',
    # 'weight',
    # 'blood_glucose',
    # # examination
    # 'examination',
    # 'general_observations',
    # 'diagnosis',
    # 'treatment',
    # 'covid_19',
    # 'referral',
    # # medicines_1
    # 'medication_1',
    # 'type_1',
    # 'dosage_1',
    # 'days_1',
    # # medicines_2
    # 'medication_2',
    # 'type_2',
    # 'dosage_2',
    # 'days_2',
    # # medicines_3
    # 'medication_3',
    # 'type_3',
    # 'dosage_3',
    # 'days_3',
    # # medicines_4
    # 'medication_4',
    # 'type_4',
    # 'dosage_4',
    # 'days_4',
    # # medicines_5
    # 'medication_5',
    # 'type_5',
    # 'dosage_5',
    # 'days_5',
    # # physiotherapy
    # 'previous_treatment',
    # 'complaint_p',
    # 'findings',
    # 'treatment_plan',
    # 'treatment_session',
    # 'recommendations',
    # 'referral',
    # 'dental_treatment',
    # 'notes',
    # 'covid_19_result',
    # # deprecated
    # 'examination_d',
    # 'medical_hx_d',
    # 'treatment_d',
    # 'diagnosis_d',
    # 'medicine_dispensed_d',
    # 'prescriptions_d',
    # 'allergies_d',

    # EMERGENCY ATTENTION SHEET
    'dateTimeEmergency',
    'namesEmergency',
    'surnamesEmergency',
    'occupationEmergency',
    'addressEmergency',
    'phoneNumberEmergency',
    'idEmergency',
    'arrivesInEmergency',
    'accidentCausesEmergency',
    'accidentPlaceEmergency',
    'weightEmergency',
    'sizeEmergency',
    'temperatureEmergency',
    'bloodPressureEmergency',
    'heartRateEmergency',
    'breathingFrequencyEmergency',
    'oxygenSaturationEmergency',
    'attentionDateHoursEmergency',
    'attendantNameEmergency',
    'notifyToNameEmergency',
    'notifyToRelationshipEmergency',
    'notifyToaddressEmergency',
    'notifyTophoneEmergency',
    'personalMedicalHistoryEmergency',
    'allergiesEmergency',
    'reasonForConsultationEmergency',
    'clinicalSummaryEmergency',
    'physicalExamEmergency',
    'evaluationEmergency',
    'diagnosisEmergency',
    'feedingEmergency',
    'medicinesEmergency',
    'examsEmergency',
    'normsOrRecommendationsEmergency',
    'destinyEmergency',
    'emergencyTypeEmergency',
    'treatingPhysicianSignatureEmergency',

    # SUBSEQUENT EVOLUTION NOTE
    'nameSurnameEvolution',
    'idEvolution',
    'communityEvolution',
    'proceedingsEvolution',
    'dateTimeEvolution',
    'bloodPressureEvolution',
    'heartRateEvolution',
    'breathingFrequencyEvolution',
    'temperatureEvolution',
    'oxygenSaturationEvolution',
    'weightEvolution',
    'sizeEvolution',
    'consultationReasonEvolution',
    'illnessHistoryEvolution',
    'objectiveEvolution',
    'evaluationEvolution',
    'diagnosisEvolution',
    'feedingEvolution',
    'medicinesEvolution',
    'examsEvolution',
    'normsOrRecommendationsEvolution',
    'treatingPhysicianEvolution',
    'minsaCodeEvolution',

    # Nursing Note
    'nameSurnameNursing',
    'idNursing',
    'communityNursing',
    'proceedingsNursing',
    'dateTimeNursing',
    'noteNursing',
    'nurseNameNursing',
    'minsaCodeNursing',

    # Ultrasound Consultation
    'dateUltrasound',
    'namesUltrasound',
    'surnameUltrasound',
    'idUltrasound',
    'originUltrasound',
    'ultrasoundPerformedUltrasound',
    'resourceUltrasound',
    'minsaCodeUltrasound',

    # Laboratory Consultation
    'dateLaboratory',
    'namesLaboratory',
    'surnameLaboratory',
    'idLaboratory',
    'originLaboratory',
    'testPerformedLaboratory',
    'resourceLaboratory',
    'minsaCodeLaboratory',

    # Odontology Consultation
    'dateOdontology',
    'namesOdontology',
    'surnameOdontology',
    'idOdontology',
    'originOdontology',
    'procedurePerformedOdontology',
    'resourceOdontology',
    'minsaCodeOdontology',

    # Family Pathological History
    'arterialHypertensionFamily',
    'diabetesMellitusFamily',
    'tuberculosisFamily',
    'cancerFamily',
    'epilepsyFamily',
    'heartDiseasesFamily',
    'nephropathiesFamily',
    'liverDiseasesFamily',
    'mentalDiseasesFamily',
    'othersFamily',
    'specifyFamily',

    # Socioeconomic Situation
    'house',
    'walls',
    'flat',
    'ceiling',
    'bedrooms',
    'animals',
    'animalsAndQuantity',
    'waterSource',
    'electricPower',
    'toilet',
    'latrine',
    'kitchen',
    'children_0_5',
    'children_6_10',
    'children_11_15',
    'adults',
    'seniors',
    'women',
    'men',
    'peopleWorkingNumber',


    # PathologicalHistory
    'childhoodDiseasesAdultPathological',
    'cardiovascularAdultPathological',
    'customizedAllergiesAdultPathological',
    'pulmonaryAdultPathological',
    'chronicDiseasesAdultPathological',
    'transfusionsAdultPathological',
    'sexuallyTransmittedInfectionsAdultPathological',
    'digestiveAdultPathological',
    'surgicalAdultPathological',
    'hospitalizationsAdultPathological',
    'othersAdultPathological',
    'specifyAdultPathological',

    # Non-Pathological History
    'alcoholismActive',
    'alcoholismAgeOfOnset',
    'alcoholismAgeOfTermination',
    'alcoholismFrequency',
    'alcoholismAmount',
    'alcoholismTypeOfLiquor',
    'smokingActive',
    'smokingStartAge',
    'smokingEndAge',
    'smokingCigarettesNumber',
    'drugsActive',
    'drugsAgeOfOnset',
    'drugsAgeOfTermination',
    'drugsFrequency',
    'drugsType',

    # Gynecological Background
    'menarche',
    'sexualLifeBeginning',
    'sexualPartnersNum',
    'familyPlanning',
    'lastMenstrualPeriodDate',
    'currentPregnancy',
    'weeksOfAmenorrhea',
    'estimatedDueDate',
    'lastDeliveryDate',
    'pregnanciesNumber',
    'births',
    'caesareanSection',
    'abortions',
    'curettage',

    # Physical Exploration
    'bloodPressurePhysicalExploration',
    'heartRatePhysicalExploration',
    'respiratoryRatePhysicalExploration',
    'temperaturePhysicalExploration',
    'oxygenSaturationPhysicalExploration',
    'weightPhysicalExploration',
    'heightPhysicalExploration',
    'bodyMassIndexPhysicalExploration',
    'consultationReasonPhysicalExploration',
    'illnessHistoryPhysicalExploration',
    'headPhysicalExploration',
    'eyesPhysicalExploration',
    'earsNoseMouthPhysicalExploration',
    'neckPhysicalExploration',
    'chestPhysicalExploration',
    'heartPhysicalExploration',
    'lungFieldsPhysicalExploration',
    'abdomenPhysicalExploration',
    'genitalsPhysicalExploration',
    'skeletalMusclePhysicalExploration',
    'extremitiesPhysicalExploration',
    'neurologicalPhysicalExploration',
    'evaluationPhysicalExploration',
    'diagnosisPhysicalExploration',
    'feedingPhysicalExploration',
    'medicinesPhysicalExploration',
    'examsPhysicalExploration',
    'normsOrRecommendationsPhysicalExploration',
    'treatingPhysicianPhysicalExploration',
    'minsaCodePhysicalExploration',

    # Postnatal History
    'apgar',
    'weightInGrams',
    'sizeInCentimeters',
    'suffocation',
    'suffocationSpecify',
    'roomingIn',
    'hospitalization',

    # Feeding
    'exclusiveBreastfeeding',
    'durationOfExclusiveBreastfeeding',
    'mixedBreastfeeding',
    'mixedBreastfeedingDuration',
    'ablactation',

    # Immunization
    'bcgDose1',
    'pentavalentDose1',
    'pentavalentDose2',
    'pentavalentDose3',
    'polioDose1',
    'polioDose2',
    'polioDose3',
    'polioReinforcement',
    'polioReinforcementDoses',
    'rotavirusDose1',
    'rotavirusDose2',
    'rotavirusDose3',
    'mmrDose1',
    'dptReinforcement',
    'dptReinforcementDoses',
    'dtDose1',
    'dtDose2',
    'dtReinforcement',
    'dtReinforcementDoses',

    # Psychomotor Development
    'suckVigorously',
    'handsClosed',
    'flexArms',
    'moorishReflex',
    'vocalize',
    'alternateLegMovement',
    'openHands',
    'socialSmile',
    'lookAtMothers',
    'followObjects',
    'reactToSound',
    'raiseHead',
    'grabObjects',
    'makeSounds',
    'holdsHead',
    'takeToy',
    'bringsObjectsToMouth',
    'locateSound',
    'turnTowardsObject',
    'playDiscover',
    'objectsPassed',
    'sitsAlone',
    'doubleSyllables',
    'imitateGestures',
    'topPliers',
    'confusingWords',
    'walkWithSupport',
    'gesturesOnDemand',
    'placeCubes',
    'sayWord',
    'walkWithoutSupport',
    'identifyTwoObjects',
    'scribbles',
    'saysThreeWords',
    'walkBackwards',
    'takesOffClothes',
    'towerFiveCubes',
    'saysShortSentences',
    'kickBall',
    'feedsItself',
    'saysLongSentences',
    'dressesAlone',
    'complexOrders',

    # Pediatric Pathological History
    'infectionsPediatricPathological',
    'chronicDiseasesPediatricPathological',
    'surgeriesPediatricPathological',
    'allergiesPediatricPathological',
    'hospitalizationsPediatricPathological',
    'othersPediatricPathological',

    # Pediatric Physical Exam
    'bloodPressurePediatricPhysical',
    'heartRatePediatricPhysical',
    'breathingFrequencyPediatricPhysical',
    'silvermanAndersonPediatricPhysical',
    'oxygenSaturationPediatricPhysical',
    'weightPediatricPhysical',
    'sizeLengthPediatricPhysical',
    'headCircunferencePediatricPhysical',
    'chestCircumferencePediatricPhysical',
    'abdominalPerimeterPediatricPhysical',
    'bodySurfaceAreaPediatricPhysical',
    'bodyMassIndexPediatricPhysical',
    'attendsNameSignPediatricPhysical',
    'consultationReasonPediatricPhysical',
    'illnessHistoryPediatricPhysical',
    'headPediatricPhysical',
    'eyesPediatricPhysical',
    'earsNoseMouthPediatricPhysical',
    'neckPediatricPhysical',
    'chestPediatricPhysical',
    'heartPediatricPhysical',
    'lungFieldsPediatricPhysical',
    'abdomenPediatricPhysical',
    'genitalsPediatricPhysical',
    'skeletalMusclePediatricPhysical',
    'extremitiesPediatricPhysical',
    'neurologicalPediatricPhysical',
    'evaluationPediatricPhysical',
    'diagnosisPediatricPhysical',
    'feedingPediatricPhysical',
    'medicinesPediatricPhysical',
    'examsPediatricPhysical',
    'normsOrRecommendationsPediatricPhysical',
    'treatingPhysicianPediatricPhysical',
    'minsaCodePediatricPhysical',
]

@dataclass
class PatientDataRow:
    # camp : str = None
    visit_date : str = None
    visit_type : str = None
    doctor : str = None
    medical_record_num : str = None
    first_name : str = None
    surname : str = None
    date_of_birth : str = None
    age : str = None
    municipality: str = None
    local_id: str = None
    address: str = None
    gender : str = None
    origin : str = None
    # hometown : str = None
    # home_country : str = None
    phone : str = None
    email : str = None
    attention_datetime : str = None
    attending_resources : str = None
    educational_status : str = None
    religion : str = None
    marital_status : str = None
    occupation : str = None
    mother_name : str = None
    father_name : str = None
    delivery_place : str = None
    delivery_datetime : str = None
    gestational_age : str = None
    delivery_care : str = None
    delivery_via : str = None
    presentation : str = None
    birthing_events : str = None

    # allergies: str = None
    # surgery_hx: str = None
    # chronic_conditions: str = None
    # current_medications: str = None
    # vaccinations: str = None
    # complaint: str = None
    # heart_rate: str = None
    # blood_pressure: str = None
    # sats: str = None
    # temp: str = None
    # respiratory_rate: str = None
    # weight: str = None
    # blood_glucose: str = None
    # examination: str = None
    # general_observations: str = None
    # diagnosis: str = None
    # treatment: str = None
    # covid_19: str = None
    # referral: str = None
    # medication_1: str = None
    # type_1: str = None
    # dosage_1: str = None
    # days_1: str = None
    # medication_2: str = None
    # type_2: str = None
    # dosage_2: str = None
    # days_2: str = None
    # medication_3: str = None
    # type_3: str = None
    # dosage_3: str = None
    # days_3: str = None
    # medication_4: str = None
    # type_4: str = None
    # dosage_4: str = None
    # days_4: str = None
    # medication_5: str = None
    # type_5: str = None
    # dosage_5: str = None
    # days_5: str = None
    # previous_treatment: str = None
    # complaint_p: str = None
    # findings: str = None
    # treatment_plan: str = None
    # treatment_session: str = None
    # recommendations: str = None
    # referral: str = None
    # dental_treatment: str = None
    # notes: str = None
    # covid_19_result: str = None
    # examination_d: str = None
    # medical_hx_d: str = None
    # treatment_d: str = None
    # diagnosis_d: str = None
    # medicine_dispensed_d: str = None
    # prescriptions_d: str = None
    # allergies_d: str = None

    # Emergency Attention Sheet
    dateTimeEmergency: str = None
    namesEmergency: str = None
    surnamesEmergency: str = None
    occupationEmergency: str = None
    addressEmergency: str = None
    phoneNumberEmergency: str = None
    idEmergency: str = None
    arrivesInEmergency: str = None
    accidentCausesEmergency: str = None
    accidentPlaceEmergency: str = None
    weightEmergency: str = None
    sizeEmergency: str = None
    temperatureEmergency: str = None
    bloodPressureEmergency: str = None
    heartRateEmergency: str = None
    breathingFrequencyEmergency: str = None
    oxygenSaturationEmergency: str = None
    attentionDateHoursEmergency: str = None
    attendantNameEmergency: str = None
    notifyToNameEmergency: str = None
    notifyToRelationshipEmergency: str = None
    notifyToaddressEmergency: str = None
    notifyTophoneEmergency: str = None
    personalMedicalHistoryEmergency: str = None
    allergiesEmergency: str = None
    reasonForConsultationEmergency: str = None
    clinicalSummaryEmergency: str = None
    physicalExamEmergency: str = None
    evaluationEmergency: str = None
    diagnosisEmergency: str = None
    feedingEmergency: str = None
    medicinesEmergency: str = None
    examsEmergency: str = None
    normsOrRecommendationsEmergency: str = None
    destinyEmergency: str = None
    emergencyTypeEmergency: str = None
    treatingPhysicianSignatureEmergency: str = None

    # SUBSEQUENT EVOLUTION NOTE
    nameSurnameEvolution: str = None
    idEvolution: str = None
    communityEvolution: str = None
    proceedingsEvolution: str = None
    dateTimeEvolution: str = None
    bloodPressureEvolution: str = None
    heartRateEvolution: str = None
    breathingFrequencyEvolution: str = None
    temperatureEvolution: str = None
    oxygenSaturationEvolution: str = None
    weightEvolution: str = None
    sizeEvolution: str = None
    consultationReasonEvolution: str = None
    illnessHistoryEvolution: str = None
    objectiveEvolution: str = None
    evaluationEvolution: str = None
    diagnosisEvolution: str = None
    feedingEvolution: str = None
    medicinesEvolution: str = None
    examsEvolution: str = None
    normsOrRecommendationsEvolution: str = None
    treatingPhysicianEvolution: str = None
    minsaCodeEvolution: str = None

    # Nursing Note
    nameSurnameNursing: str = None
    idNursing: str = None
    communityNursing: str = None
    proceedingsNursing: str = None
    dateTimeNursing: str = None
    noteNursing: str = None
    nurseNameNursing: str = None
    minsaCodeNursing: str = None

    # Ultrasound Consultation
    dateUltrasound: str = None
    namesUltrasound: str = None
    surnameUltrasound: str = None
    idUltrasound: str = None
    originUltrasound: str = None
    ultrasoundPerformedUltrasound: str = None
    resourceUltrasound: str = None
    minsaCodeUltrasound: str = None

    # Laboratory Consultation
    dateLaboratory: str = None
    namesLaboratory: str = None
    surnameLaboratory: str = None
    idLaboratory: str = None
    originLaboratory: str = None
    testPerformedLaboratory: str = None
    resourceLaboratory: str = None
    minsaCodeLaboratory: str = None

    # Odontology Consultation
    dateOdontology: str = None
    namesOdontology: str = None
    surnameOdontology: str = None
    idOdontology: str = None
    originOdontology: str = None
    procedurePerformedOdontology: str = None
    resourceOdontology: str = None
    minsaCodeOdontology: str = None

    # Family Pathological History
    arterialHypertensionFamily:str = None
    diabetesMellitusFamily:str = None
    tuberculosisFamily:str = None
    cancerFamily:str = None
    epilepsyFamily:str = None
    heartDiseasesFamily:str = None
    nephropathiesFamily:str = None
    liverDiseasesFamily:str = None
    mentalDiseasesFamily:str = None
    othersFamily:str = None
    specifyFamily:str = None

    # Socioeconomic Situation
    house:str = None
    walls:str = None
    flat:str = None
    ceiling:str = None
    bedrooms:str = None
    animals:str = None
    animalsAndQuantity:str = None
    waterSource:str = None
    electricPower:str = None
    toilet:str = None
    latrine:str = None
    kitchen:str = None
    children_0_5:str = None
    children_6_10:str = None
    children_11_15:str = None
    adults:str = None
    seniors:str = None
    women:str = None
    men:str = None
    peopleWorkingNumber:str = None


    # Pathological History
    childhoodDiseasesAdultPathological : str = None
    cardiovascularAdultPathological : str = None
    customizedAllergiesAdultPathological : str = None
    pulmonaryAdultPathological : str = None
    chronicDiseasesAdultPathological : str = None
    transfusionsAdultPathological : str = None
    sexuallyTransmittedInfectionsAdultPathological : str = None
    digestiveAdultPathological : str = None
    surgicalAdultPathological : str = None
    hospitalizationsAdultPathological : str = None
    othersAdultPathological : str = None
    specifyAdultPathological : str = None


    # Non-Pathological
    alcoholismActive: str = None
    alcoholismAgeOfOnset: str = None
    alcoholismAgeOfTermination: str = None
    alcoholismFrequency: str = None
    alcoholismAmount: str = None
    alcoholismTypeOfLiquor: str = None
    smokingActive: str = None
    smokingStartAge: str = None
    smokingEndAge: str = None
    smokingCigarettesNumber: str = None
    drugsActive: str = None
    drugsAgeOfOnset: str = None
    drugsAgeOfTermination: str = None
    drugsFrequency: str = None
    drugsType: str = None

    # Gynecological Background
    menarche:str = None
    sexualLifeBeginning:str = None
    sexualPartnersNum:str = None
    familyPlanning:str = None
    lastMenstrualPeriodDate:str = None
    currentPregnancy:str = None
    weeksOfAmenorrhea:str = None
    estimatedDueDate:str = None
    lastDeliveryDate:str = None
    pregnanciesNumber:str = None
    births:str = None
    caesareanSection:str = None
    abortions:str = None
    curettage:str = None

    # Physical Exploration
    bloodPressurePhysicalExploration: str = None
    heartRatePhysicalExploration: str = None
    respiratoryRatePhysicalExploration: str = None
    temperaturePhysicalExploration: str = None
    oxygenSaturationPhysicalExploration: str = None
    weightPhysicalExploration: str = None
    heightPhysicalExploration: str = None
    bodyMassIndexPhysicalExploration: str = None
    consultationReasonPhysicalExploration: str = None
    illnessHistoryPhysicalExploration: str = None
    headPhysicalExploration: str = None
    eyesPhysicalExploration: str = None
    earsNoseMouthPhysicalExploration: str = None
    neckPhysicalExploration: str = None
    chestPhysicalExploration: str = None
    heartPhysicalExploration: str = None
    lungFieldsPhysicalExploration: str = None
    abdomenPhysicalExploration: str = None
    genitalsPhysicalExploration: str = None
    skeletalMusclePhysicalExploration: str = None
    extremitiesPhysicalExploration: str = None
    neurologicalPhysicalExploration: str = None
    evaluationPhysicalExploration: str = None
    diagnosisPhysicalExploration: str = None
    feedingPhysicalExploration: str = None
    medicinesPhysicalExploration: str = None
    examsPhysicalExploration: str = None
    normsOrRecommendationsPhysicalExploration: str = None
    treatingPhysicianPhysicalExploration: str = None
    minsaCodePhysicalExploration: str = None

    # Adult Immunizations
    pneumococcusAdultImmunizations: str = None
    influenzaAdultImmunizations: str = None
    tetanusAdultImmunizations: str = None
    hepatitisBAdultImmunizations: str = None
    covidAdultImmunizations: str = None
    tbAdultImmunizations: str = None
    othersAdultImmunizations: str = None
    # Postatal History
    apgar: str = None
    weightInGrams: str = None
    sizeInCentimeters: str = None
    suffocation: str = None
    suffocationSpecify: str = None
    roomingIn: str = None
    hospitalization: str = None

    # Feeding
    exclusiveBreastfeeding: str = None
    durationOfExclusiveBreastfeeding: str = None
    mixedBreastfeeding: str = None
    mixedBreastfeedingDuration: str = None
    ablactation: str = None

    # Immunization
    bcgDose1: str = None
    pentavalentDose1: str = None
    pentavalentDose2: str = None
    pentavalentDose3: str = None
    polioDose1: str = None
    polioDose2: str = None
    polioDose3: str = None
    polioReinforcement: str = None
    polioReinforcementDoses: str = None
    rotavirusDose1: str = None
    rotavirusDose2: str = None
    rotavirusDose3: str = None
    mmrDose1: str = None
    dptReinforcement: str = None
    dptReinforcementDoses: str = None
    dtDose1: str = None
    dtDose2: str = None
    dtReinforcement: str = None
    dtReinforcementDoses: str = None

    # Psychomotor Development
    suckVigorously: str = None
    handsClosed: str = None
    flexArms: str = None
    moorishReflex: str = None
    vocalize: str = None
    alternateLegMovement: str = None
    openHands: str = None
    socialSmile: str = None
    lookAtMothers: str = None
    followObjects: str = None
    reactToSound: str = None
    raiseHead: str = None
    grabObjects: str = None
    makeSounds: str = None
    holdsHead: str = None
    takeToy: str = None
    bringsObjectsToMouth: str = None
    locateSound: str = None
    turnTowardsObject: str = None
    playDiscover: str = None
    objectsPassed: str = None
    sitsAlone: str = None
    doubleSyllables: str = None
    imitateGestures: str = None
    topPliers: str = None
    confusingWords: str = None
    walkWithSupport: str = None
    gesturesOnDemand: str = None
    placeCubes: str = None
    sayWord: str = None
    walkWithoutSupport: str = None
    identifyTwoObjects: str = None
    scribbles: str = None
    saysThreeWords: str = None
    walkBackwards: str = None
    takesOffClothes: str = None
    towerFiveCubes: str = None
    saysShortSentences: str = None
    kickBall: str = None
    feedsItself: str = None
    saysLongSentences: str = None
    dressesAlone: str = None
    complexOrders: str = None

    # Pediatric Pathological History
    infections: str = None
    chronicDiseases: str = None
    surgeries: str = None
    allergies: str = None
    hospitalizations: str = None
    others: str = None

    # Pediatric Physical Exam
    bloodPressurePediatricPhysical: str = None
    heartRatePediatricPhysical: str = None
    breathingFrequencyPediatricPhysical: str = None
    silvermanAndersonPediatricPhysical: str = None
    oxygenSaturationPediatricPhysical: str = None
    weightPediatricPhysical: str = None
    sizeLengthPediatricPhysical: str = None
    headCircunferencePediatricPhysical: str = None
    chestCircumferencePediatricPhysical: str = None
    abdominalPerimeterPediatricPhysical: str = None
    bodySurfaceAreaPediatricPhysical: str = None
    bodyMassIndexPediatricPhysical: str = None
    attendsNameSignPediatricPhysical: str = None
    consultationReasonPediatricPhysical: str = None
    illnessHistoryPediatricPhysical: str = None
    headPediatricPhysical: str = None
    eyesPediatricPhysical: str = None
    earsNoseMouthPediatricPhysical: str = None
    neckPediatricPhysical: str = None
    chestPediatricPhysical: str = None
    heartPediatricPhysical: str = None
    lungFieldsPediatricPhysical: str = None
    abdomenPediatricPhysical: str = None
    genitalsPediatricPhysical: str = None
    skeletalMusclePediatricPhysical: str = None
    extremitiesPediatricPhysical: str = None
    neurologicalPediatricPhysical: str = None
    evaluationPediatricPhysical: str = None
    diagnosisPediatricPhysical: str = None
    feedingPediatricPhysical: str = None
    medicinesPediatricPhysical: str = None
    examsPediatricPhysical: str = None
    normsOrRecommendationsPediatricPhysical: str = None
    treatingPhysicianPediatricPhysical: str = None
    minsaCodePediatricPhysical: str = None


# COLUMN_TYPES = [str, None, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, float, str,
#                 float, float, float, float, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str,
#                 str, str]


# class PatientDataImporter:
#     def __init__(self, data_file: FileStorage):
#         self.data_filename = self._write_file_to_tempfile(data_file)

#     def run(self):
#         all_rows = [self._parse_row(row) for row in self.iter_data_rows()]
#         print('Creating patients...')
#         self._create_patients(all_rows)
#         print('Creating visits...')
#         self._create_visits(all_rows)

#     def _parse_row(self, row):
#         if len(row) != 41:
#             raise WebError('All data rows must have exactly 41 data points.', 400)
#         values = [self._parse_cell(value, formatter) for value, formatter in zip(row, COLUMN_TYPES)]
#         return PatientDataRow(**dict(zip(COLUMNS, values)))

#     def _parse_cell(self, cell, formatter):
#         if cell == 'Nil' or cell is None:
#             return None
#         if formatter is None:
#             return cell
#         return formatter(cell)

#     @staticmethod
#     def _write_file_to_tempfile(data_file: FileStorage):
#         handle = NamedTemporaryFile('wb', delete=False, suffix='.xlsx')
#         data_file.save(handle)
#         handle.close()
#         print('Upload written to', handle.name)
#         return handle.name

#     def iter_data_rows(self):
#         wb = load_workbook(self.data_filename)
#         ws = wb.active
#         for idx, row in enumerate(ws.iter_rows(min_row=3, max_col=41, values_only=True)):
#             if all(x is None for x in row):
#                 continue
#             yield row

#     def _create_patients(self, rows: Iterable[PatientDataRow]):
#         for patient_data in set(map(lambda r: (r.first_name, r.surname, r.gender, r.home_country, r.age), rows)):
#             first_name, surname, gender, home_country, age = patient_data
#             if not patient_from_key_data(first_name, surname, home_country, self._parse_sex(gender)):
#                 self._create_patient(first_name, surname, home_country, gender, age)

#     def _create_patient(self, given_name, surname, home_country, sex, age):
#         given_name_ls = LanguageString(id=str(uuid.uuid4()), content_by_language={'en': given_name})
#         surname_ls = LanguageString(id=str(uuid.uuid4()), content_by_language={'en': surname})
#         inferred_dob = self._infer_dob(age)
#         patient = Patient(
#             id=str(uuid.uuid4()),
#             edited_at=datetime.now(),
#             given_name=given_name_ls,
#             surname=surname_ls,
#             date_of_birth=inferred_dob,
#             sex=self._parse_sex(sex),
#             country=LanguageString(id=str(uuid.uuid4()), content_by_language={'en': home_country}),
#             phone=None,
#             hometown=None
#         )
#         add_patient(patient)

#     @staticmethod
#     def _parse_sex(sex_str):
#         if sex_str is None:
#             return None
#         elif 'm' in sex_str.lower():
#             return 'M'
#         elif 'f' in sex_str.lower():
#             return 'F'
#         else:
#             return None

#     def _infer_dob(self, age_string):
#         try:
#             int_prefix = int(''.join(itertools.takewhile(str.isnumeric, age_string)))
#             today = date.today()
#             if 'months' in age_string:
#                 return today - timedelta(days=30 * int_prefix)
#             elif 'weeks' in age_string:
#                 return today - timedelta(weeks=int_prefix)
#             elif 'days' in age_string:
#                 return today - timedelta(days=int_prefix)
#             else:
#                 # Assume years if no unit is specified
#                 return today - timedelta(days=365 * int_prefix)
#         except (ValueError, TypeError):
#             return date(1900, 1, 1)

#     @staticmethod
#     def _parse_date(date_str):
#         if isinstance(date_str, date) or isinstance(date_str, datetime):
#             return date_str
#         try:
#             dt = pd.to_datetime(date_str, dayfirst=True).to_pydatetime()
#             return date(year=dt.year, month=dt.month, day=dt.day)
#         except dateutil.parser._parser.ParserError:
#             return None

#     def _create_visits(self, rows: Iterable[PatientDataRow]):
#         for row in rows:
#             patient_id = patient_from_key_data(row.first_name, row.surname, row.home_country, self._parse_sex(row.gender))
#             if not patient_id:
#                 print('Warning: unknown patient; skipping.')
#                 continue
#             visit_date = self._parse_date(row.visit_date)
#             visit_id, visit_timestamp = first_visit_by_patient_and_date(patient_id, visit_date)

#             # TODO: The data import format does not currently specify a clinic. Since
#             # current Hikma instances are single clinic anyway, just get the most common
#             # clinic (in case there is a demo one with few if any visits) and use that.
#             clinic_id = get_most_common_clinic()

#             # TODO: The data import format does not currently specify a provider in a format
#             # that we can use. So for now, use a per-instance default provider that is set via
#             # environment variable.
#             provider_id = DEFAULT_PROVIDER_ID_FOR_IMPORT

#             if visit_id is None:
#                 visit_id = str(uuid.uuid4())
#                 visit_timestamp = datetime.combine(visit_date, datetime.min.time())
#                 visit = Visit(
#                     id=visit_id,
#                     patient_id=patient_id,
#                     edited_at=datetime.now(),
#                     clinic_id=clinic_id,
#                     provider_id=provider_id,
#                     check_in_timestamp=visit_timestamp
#                 )
#                 add_visit(visit)

#                 # Until we implement full deletion, only add visit the first time it is seen.
#                 self._update_events(patient_id, visit_id, visit_timestamp, row)

#     def _update_events(self, patient_id: str, visit_id: str, visit_timestamp: datetime, row: PatientDataRow):
#         # TODO: This will need to be replaced with a mode of deletion that persists through synchronization.
#         # clear_all_events(visit_id)
#         if row.allergies:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Allergies', row.allergies)
#         if any([row.dispensed_medicine_1, row.dispensed_medicine_2,
#                 row.dispensed_medicine_3, row.dispensed_medicine_4]):
#             self._add_dispensed_medicine_event(patient_id, visit_id, visit_timestamp, row)
#         if row.presenting_complaint:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Complaint', row.presenting_complaint)
#         if any([row.heart_rate, row.blood_pressure, row.o2_sats,
#                 row.respiratory_rate, row.temperature, row.blood_glucose]):
#             self._add_vitals_event(patient_id, visit_id, visit_timestamp, row)
#         if row.examination:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Examination', row.examination)
#         if row.diagnosis:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Diagnosis', row.diagnosis)
#         if row.treatment:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Treatment', row.treatment)
#         if row.prescription:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Prescriptions', row.prescription)
#         if row.notes:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Notes', row.notes)
#         if row.camp:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Camp', row.camp)

#     def _add_text_event(self, patient_id: str, visit_id: str, visit_timestamp: datetime,
#                         event_type: str, event_metadata: str):
#         event = Event(
#             id=str(uuid.uuid4()),
#             patient_id=patient_id,
#             visit_id=visit_id,
#             event_type=event_type,
#             event_timestamp=visit_timestamp,
#             event_metadata=event_metadata,
#             edited_at=datetime.now(),
#         )
#         add_event(event)

#     def _add_dispensed_medicine_event(self,  patient_id: str, visit_id: str, visit_timestamp: datetime, row: PatientDataRow):
#         data = [
#             (row.dispensed_medicine_1, row.dispensed_medicine_quantity_1),
#             (row.dispensed_medicine_2, row.dispensed_medicine_quantity_2),
#             (row.dispensed_medicine_3, row.dispensed_medicine_quantity_3),
#             (row.dispensed_medicine_4, row.dispensed_medicine_quantity_4),
#         ]
#         content = '\n'.join([': '.join(r) for r in data if all(r)])
#         event = Event(
#             id=str(uuid.uuid4()),
#             patient_id=patient_id,
#             visit_id=visit_id,
#             event_type='Medicine Dispensed',
#             event_timestamp=visit_timestamp,
#             event_metadata=content,
#             edited_at=datetime.now(),
#         )
#         add_event(event)

#     def _add_vitals_event(self,  patient_id: str, visit_id: str, visit_timestamp: datetime, row: PatientDataRow):
#         try:
#             diastolic, systolic = row.blood_pressure.split('/')
#         except (ValueError, AttributeError):
#             diastolic = None
#             systolic = None

#         data = {
#             'heartRate': as_string(row.heart_rate),
#             'systolic': as_string(systolic),
#             'diastolic': as_string(diastolic),
#             'sats': as_string(row.o2_sats),
#             'temp': as_string(row.temperature),
#             'respiratoryRate': as_string(row.respiratory_rate),
#             'bloodGlucose': as_string(row.blood_glucose)
#         }

#         event = Event(
#             id=str(uuid.uuid4()),
#             patient_id=patient_id,
#             visit_id=visit_id,
#             event_type='Vitals',
#             event_timestamp=visit_timestamp,
#             event_metadata=json.dumps(data),
#             edited_at=datetime.now(),
#         )
#         add_event(event)
